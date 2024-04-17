import openpyxl

def criar_excel_dados(lista_dados):
    try:
        wb = openpyxl.Workbook()
        planilha = wb.active

        cabecalhos = list(lista_dados[0].keys())
        for coluna, cabecalho in enumerate(cabecalhos, start=1):
            planilha.cell(row=1, column=coluna, value=cabecalho)


        for linha, item in enumerate(lista_dados, start=2):
            for coluna, valor in enumerate(item.values(), start=1):
                planilha.cell(row=linha, column=coluna, value=valor)

        wb.save('dados_excel.xlsx')
        
        return True
    except Exception as e:
        print(f'Erro metodo criar_excel_dados: {e}')